from bs4 import BeautifulSoup
import requests
from datetime import datetime
import pprint
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import os
import time
import re


def table_decomposition(url):
    try:
        response = requests.get(url)
        if response.status_code != 200:
            print(f"Ошибка при запросе {url}. Статус код: {response.status_code}")
            return []

        html_content = response.text
        soup = BeautifulSoup(html_content, 'html.parser')

        # Извлекаем заголовок страницы
        page_title = soup.find('h1').get_text(strip=True)

        # Находим все таблицы на странице
        tables = soup.find_all("table")
        print(f"Найдено таблиц: {len(tables)}")

        # Инициализируем список для хранения записей
        table_data = []

        # Проходим по всем таблицам
        for table_index, table in enumerate(tables):
            # Проверяем, содержит ли первая строка таблицы класс "min-hidden"
            first_row = table.find('tr', class_='min-hidden')
            if first_row is not None:
                # Извлекаем заголовки столбцов
                headers = [th.get_text(strip=True).lower() for th in first_row.find_all('th')]

                # Проходим по строкам таблицы (начиная со 2-й строки данных)
                rows = table.find_all('tr')[2:]

                for row in rows:
                    cells = row.find_all('td')

                    # Проверяем, что количество ячеек совпадает с количеством заголовков
                    if len(cells) == len(headers):
                        # Инициализируем переменные для необходимых данных
                        article = None
                        name = page_title  # Название страницы используется как имя по умолчанию
                        price = None

                        # Проходим по каждой ячейке и соответствующему заголовку
                        for i, cell in enumerate(cells):
                            header = headers[i]

                            if 'артикул' in header:
                                article = cell.get_text(strip=True)
                            elif 'цена' in header:
                                # Извлекаем только цифры из цены
                                price_text = cell.get_text(strip=True)
                                price = ''.join(re.findall(r'\d+', price_text))
                            #else:
                        # Любой другой заголовок пока не используем (можно добавить при необходимости)

                        # Формируем строку данных только если есть артикул и цена
                        if article:
                            row_data = {
                                'Артикул': article,
                                'Название': name,
                                'Цена': price,
                                'Ссылка': url,
                                'Дата': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            }
                            table_data.append(row_data)

            else:
                print(f"Пропускаем таблицу без класса 'min-hidden'")

        # Сохраняем данные в Excel, если таблицы были найдены
        if table_data:
            pprint.pprint(table_data)
            print_to_excel(table_data)

        return table_data

    except requests.exceptions.RequestException as e:
        print(f"Ошибка запроса: {e}")
        return []


def print_to_excel(table_data):
    #print("print_to_excel")

    # Преобразование строковых значений цен в числа с плавающей запятой и округление до 2 знаков
    for row in table_data:
        price_key = next((key for key in row if 'цена' in key.lower()), None)
        if price_key and row[price_key].replace('.', '', 1).isdigit():  # Проверка, что цена состоит из цифр
            row[price_key] = round(float(row[price_key]), 2)  # Преобразуем в float и округляем до 2 знаков

    # Создание имени файла с текущей датой
    filename = f"competitors_parsing_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

    sheet_name = f"valtec_ru_{datetime.now().strftime('%Y-%m-%d')}"

    # Проверяем, существует ли файл
    if os.path.exists(filename):
        # Если файл существует, открываем его
        wb = load_workbook(filename)

        # Проверяем, есть ли лист с нужным именем
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            # Если листа нет, создаём новый
            ws = wb.create_sheet(title=sheet_name)
            # Запись заголовков столбцов (это ключи первого словаря)
            headers = list(table_data[0].keys())
            ws.append(headers)

    else:
        # Если файл не существует, создаём новый
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        # Запись заголовков столбцов (это ключи первого словаря)
        headers = list(table_data[0].keys())
        ws.append(headers)

    # Определяем, с какой строки начинать добавление данных
    start_row = ws.max_row + 1

    # Запись данных
    for row in table_data:
        ws.append(list(row.values()))  # добавляем значения каждой строки

    # Сохранение файла
    wb.save(filename)

    print(f"Данные о {len(table_data)} товарах успешно сохранены в файл {filename}.")


def get_urls_by_classes(url, class_names):
    try:
        response = requests.get(url)
        if response.status_code != 200:  # Проверяем код ответа
            print(f"Ошибка при запросе {url}. Статус код: {response.status_code}")
            return []

        html = response.text
        soup = BeautifulSoup(html, "html.parser")
        full_links = []

        # Ищем div с любым из указанных классов
        catalog_div = None
        for class_name in class_names:
            catalog_div = soup.find("div", class_=class_name)
            if catalog_div:  # Если нашли блок с этим классом, выходим из цикла
                break

        # Если блок найден, извлекаем ссылки
        if catalog_div:
            links = catalog_div.find_all("a")
            # Собираем полные ссылки
            for link in links:
                full_link = "https://valtec.ru" + str(link.get('href'))
                full_links.append(full_link)

        return full_links

    except requests.exceptions.RequestException as e:  # Ловим любые ошибки с запросом
        print(f"Ошибка запроса: {e}")
        return []


def recursive_parse(url, class_names, total_products_found):
    """Рекурсивная функция для обработки вложенных разделов каталога."""
    # Получаем все ссылки на странице
    links = get_urls_by_classes(url, class_names)

    if links:
        for link in links:
            # Проверяем, есть ли вложенные разделы на следующем уровне
            nested_links = get_urls_by_classes(link, class_names)

            if nested_links:
                # Если есть вложенные разделы, продолжаем углубляться рекурсивно
                total_products_found = recursive_parse(link, class_names, total_products_found)
            else:
                # Если нет вложенных разделов, вызываем парсинг таблицы
                table_data = table_decomposition(link)
                total_products_found += len(table_data)  # Увеличиваем счётчик товаров
    return total_products_found


# Основной цикл программы
def main():
    url0 = "https://valtec.ru/catalog/"
    class_names = ["catalog2 index test", "catalog3 group", "catalog-2 container", "catalog2 subgroup", "catalog2 index"]

    # Начало отсчёта времени
    start_time = time.time()

    # Счётчик для подсчёта количества найденных товаров
    total_products_found = 0

    # Рекурсивный вызов для парсинга корневого каталога
    total_products_found = recursive_parse(url0, class_names, total_products_found)

    # Конец отсчёта времени
    end_time = time.time()

    # Подсчитываем потраченное время
    elapsed_time = end_time - start_time
    elapsed_time_minutes = elapsed_time / 60  # переводим в минуты

    # Вывод результатов
    print(f"Программа завершена.")
    print(f"Найдено товаров: {total_products_found}")
    print(f"Время выполнения программы: {elapsed_time:.2f} секунд ({elapsed_time_minutes:.2f} минут)")


if __name__ == "__main__":
    main()