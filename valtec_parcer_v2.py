from bs4 import BeautifulSoup
import requests
from datetime import datetime
import pprint
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import os
import time
import re


def table_decomposition(url):
    try:
        response = requests.get(url)
        if response.status_code != 200:  # Проверяем код ответа
            print(f"Ошибка при запросе {url}. Статус код: {response.status_code}")
            return []

        html_content = response.text
        # Создание объекта BeautifulSoup
        soup = BeautifulSoup(html_content, 'html.parser')

        # Извлекаем заголовок страницы (тег <h1>)
        page_title = soup.find('h1').get_text(strip=True)

        # Находим все таблицы на странице
        tables = soup.find_all("table")
        print(f"Найдено таблиц: {len(tables)}")

        # Инициализируем список для хранения записей
        table_data = []

        # Проходим по всем таблицам
        for table_index, table in enumerate(tables):
            # Проверяем, содержит ли первая строка таблицы класс "min-hidden"
            first_row = table.find('tr', class_='min-hidden')
            if first_row is not None:
                # Дополнительно проверяем, содержат ли заголовки ключевые слова (например, "Артикул", "Размер", "Цена")
                headers = [th.get_text(strip=True) for th in first_row.find_all('th')]

                if first_row or any(keyword in headers for keyword in ["артикул", "размер", "цена"]):
                    # Проходим по строкам таблицы (начиная со 2-й строки данных)
                    rows = table.find_all('tr')[2:]

                    for row_index, row in enumerate(rows):
                        # Извлекаем все ячейки в строке
                        cells = row.find_all('td')

                        # Проверяем, что количество ячеек совпадает с количеством заголовков
                        if len(cells) == len(headers):
                            # Создаем запись (словарь) для каждой строки таблицы
                            row_data = {
                                'Название': page_title  # Сначала добавляем заголовок страницы
                            }
                            for i, cell in enumerate(cells):
                                # Если заголовок содержит "Цена", обрабатываем как цену
                                if 'цена' in headers[i].lower():
                                    # Проверяем, что именно содержится в ячейке с ценой
                                    price_text = cell.get_text(strip=True)

                                    # Используем регулярное выражение для извлечения всех цифр (включая дробные)
                                    price_text = ''.join(re.findall(r'\d+', price_text))  # Извлекаем только цифры

                                    row_data[headers[i]] = price_text
                                else:
                                    # Добавляем текст в соответствующий заголовок
                                    row_data[headers[i]] = cell.get_text(strip=True)

                            # Добавляем ссылку на товар в последнюю очередь
                            row_data['Ссылка'] = url
                            # Добавляем запись в список
                            table_data.append(row_data)
            else:
                print(f"Пропускаем таблицу без класса 'min-hidden'")

        # Сохраняем данные в Excel, если таблицы были найдены
        if table_data:
            pprint.pprint(table_data)
            print_to_excel(table_data)

        return table_data

    except requests.exceptions.RequestException as e:  # Ловим любые ошибки с запросом
        print(f"Ошибка запроса: {e}")
        return []


def print_to_excel(table_data):
    print("print_to_excel")
    # Преобразование строковых значений цен в числа с плавающей запятой и округление до 2 знаков
    for row in table_data:
        # Определяем, какой заголовок относится к цене
        price_key = next((key for key in row if 'цена' in key.lower()), None)
        if price_key and row[price_key].isdigit():  # Проверка, что цена состоит из цифр
            # Преобразуем значение цены в float и округляем до 2 знаков
            row[price_key] = round(float(row[price_key]), 2)

    # Получаем текущую дату и время в формате 'YYYY-MM-DD HH:MM:SS'
    current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Добавляем текущую дату и время в каждый ряд данных
    for row in table_data:
        row['Дата'] = current_datetime

    # Создание имени файла с текущей датой
    filename = f"competitors_parsing_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

    # Проверяем, существует ли файл
    if os.path.exists(filename):
        # Если файл существует, открываем его для добавления данных
        wb = load_workbook(filename)
        ws = wb.active
    else:
        # Если файл не существует, создаём новый
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"valtec_ru_{datetime.now().strftime('%Y-%m-%d')}"
        # Запись заголовков столбцов (это ключи первого словаря)
        headers = list(table_data[0].keys())  # теперь включая "Дата" и "Ссылка"
        ws.append(headers)

    # Определяем, с какой строки начинать добавление данных
    start_row = ws.max_row + 1

    # Запись данных
    for row in table_data:
        ws.append(list(row.values()))  # добавляем значения каждой строки

    # Сохранение файла
    wb.save(filename)

    print(f"Данные успешно сохранены в файл {filename}.")

''' ======блок с поуровневой обработкой разделов каталога - из 1 версии парсера ======
def get_urls_lvl_2(url):
    try:
        response = requests.get(url)
        if response.status_code != 200:  # Проверяем код ответа
            print(f"Ошибка при запросе {url}. Статус код: {response.status_code}")
            return []

        html = response.text
        soup = BeautifulSoup(html, "html.parser")
        full_links2 = []
        # Находим div с классом "catalog2 subgroup"
        catalog_div = soup.find("div", class_="catalog2 subgroup")

        # Извлекаем все ссылки (теги <a>) внутри этого блока
        if catalog_div:
            links2 = catalog_div.find_all("a")
            # Собираем полные ссылки
            for link in links2:
                full_link2 = "https://valtec.ru" + str(link.get('href'))
                full_links2.append(full_link2)

        return full_links2

    except requests.exceptions.RequestException as e:  # Ловим любые ошибки с запросом
        print(f"Ошибка запроса: {e}")
        return []


def get_urls_lvl_1(url):
    try:
        response = requests.get(url)
        if response.status_code != 200:  # Проверяем код ответа
            print(f"Ошибка при запросе {url}. Статус код: {response.status_code}")
            return []

        html = response.text
        soup = BeautifulSoup(html, "html.parser")
        full_links1 = []

        # Находим родительский блок с классом "catalog-2 container"
        catalog_div = soup.find("div", class_="catalog-2 container")

        # Если такой блок найден
        if catalog_div:
            # Находим все элементы div с классом "col" внутри этого блока
            cols = catalog_div.find_all("div", class_="col")

            # Извлекаем ссылки <a> из каждого элемента "col"
            for col in cols:
                link = col.find("a")
                if link and link.get('href'):
                    full_link1 = "https://valtec.ru" + str(link.get('href'))
                    full_links1.append(full_link1)

        return full_links1

    except requests.exceptions.RequestException as e:  # Ловим любые ошибки с запросом
        print(f"Ошибка запроса: {e}")
        return []


def get_urls_lvl_0(url):
    try:
        response = requests.get(url)
        if response.status_code != 200:  # Проверяем код ответа, если не 200 - пропускаем
            print(f"Ошибка при запросе {url}. Статус код: {response.status_code}")
            return []

        html = response.text
        soup = BeautifulSoup(html, "html.parser")
        full_links0 = []
        # Находим div с классом "catalog2 index test"
        catalog_div = soup.find("div", class_="catalog2 index test")

        # Извлекаем все ссылки (теги <a>) внутри этого блока
        if catalog_div:
            links0 = catalog_div.find_all("a")
            # Собираем полные ссылки
            for link in links0:
                full_link0 = "https://valtec.ru" + str(link.get('href'))
                full_links0.append(full_link0)

        return full_links0

    except requests.exceptions.RequestException as e:  # Ловим любые ошибки с запросом
        print(f"Ошибка запроса: {e}")
        return []  # Возвращаем пустой список, чтобы продолжить выполнение программы

'''
def get_urls_by_classes(url, class_names):
    try:
        response = requests.get(url)
        if response.status_code != 200:  # Проверяем код ответа
            print(f"Ошибка при запросе {url}. Статус код: {response.status_code}")
            return []

        html = response.text
        soup = BeautifulSoup(html, "html.parser")
        full_links = []

        # Ищем div с любым из указанных классов
        catalog_div = None
        for class_name in class_names:
            catalog_div = soup.find("div", class_=class_name)
            if catalog_div:  # Если нашли блок с этим классом, выходим из цикла
                break

        # Если блок найден, извлекаем ссылки
        if catalog_div:
            links = catalog_div.find_all("a")
            # Собираем полные ссылки
            for link in links:
                full_link = "https://valtec.ru" + str(link.get('href'))
                full_links.append(full_link)

        return full_links

    except requests.exceptions.RequestException as e:  # Ловим любые ошибки с запросом
        print(f"Ошибка запроса: {e}")
        return []




url0 = "https://valtec.ru/catalog/"

class_names = ["catalog2 index test", "catalog3 group","catalog-2 container", "catalog2 subgroup"]
# Начало отсчёта времени
start_time = time.time()

# Счётчик для подсчёта количества найденных товаров
total_products_found = 0

'''здесь в параметрах функции передаются возможные варианты классов, в которых находятся разделы каталогов'''
# Основной цикл парсинга
for x in get_urls_by_classes(url0, class_names):  # парсим корневой уровень каталога
    print(f"Прогон цикла ссылок 0 уровня: {x}")
    for i in get_urls_by_classes(x, class_names):  # парсим первый уровень каталога
        print(f"Прогон цикла ссылок 1 уровня: {i}")
        for j in get_urls_by_classes(i, class_names):  # парсим второй уровень каталога, там уже лежат карточки товаров
            print(f"Прогон цикла ссылок 2 уровня: {j}")
            #table_decomposition(j)
            # Вызываем функцию парсинга таблицы и сохраняем результат
            table_data = table_decomposition(j)

            # Увеличиваем счётчик товаров на количество найденных карточек товаров
            total_products_found += len(table_data)

# Конец отсчёта времени
end_time = time.time()

# Подсчитываем потраченное время
elapsed_time = end_time - start_time
elapsed_time_minutes = elapsed_time / 60  # переводим в минуты

# Вывод результатов
print(f"Программа завершена.")
print(f"Найдено товаров: {total_products_found}")
print(f"Время выполнения программы: {elapsed_time:.2f} секунд ({elapsed_time_minutes:.2f} минут)")