""" ========================== Stout.ru ==========================
Корневая папка каталога - https://www.stout.ru/catalog/
Все ссылки на подразделы имеют класс "but but-1"
Уже на первом уровне каталога можно раскрутить все внутренние товары через пагинацию.
В карточку проваливаться не нужно, артикул, название и цену можно взять в списке товаров.
==================================================================
"""

from bs4 import BeautifulSoup
import requests
from datetime import datetime
import pprint
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import os
import time
import re


def print_to_excel(table_data):
    #print("print_to_excel")

    # Преобразование строковых значений цен в числа с плавающей запятой и округление до 2 знаков
    for row in table_data:
        price_key = next((key for key in row if 'цена' in key.lower()), None)
        if price_key and row[price_key].replace('.', '', 1).isdigit():  # Проверка, что цена состоит из цифр
            row[price_key] = round(float(row[price_key]), 2)  # Преобразуем в float и округляем до 2 знаков

    # Создание имени файла с текущей датой
    filename = f"competitors_parsing_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

    sheet_name = f"stout_ru_{datetime.now().strftime('%Y-%m-%d')}"

    # Проверяем, существует ли файл
    if os.path.exists(filename):
        # Если файл существует, открываем его
        wb = load_workbook(filename)

        # Проверяем, есть ли лист с нужным именем
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            # Если листа нет, создаём новый
            ws = wb.create_sheet(title=sheet_name)
            # Запись заголовков столбцов (это ключи первого словаря)
            headers = list(table_data[0].keys())
            ws.append(headers)

    else:
        # Если файл не существует, создаём новый
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        # Запись заголовков столбцов (это ключи первого словаря)
        headers = list(table_data[0].keys())
        ws.append(headers)

    # Определяем, с какой строки начинать добавление данных
    start_row = ws.max_row + 1

    # Запись данных
    for row in table_data:
        ws.append(list(row.values()))  # добавляем значения каждой строки

    # Сохранение файла
    wb.save(filename)

    print(f"Данные о {len(table_data)} товарах успешно сохранены в файл {filename}.")



def get_urls_lvl_0(url): # Разбираем корневой уровень каталога, получаем ссылки на разделы первого уровня
    try:
        response = requests.get(url)
        if response.status_code != 200:  # Проверяем код ответа, если не 200 - пропускаем
            print(f"Ошибка при запросе {url}. Статус код: {response.status_code}")
            return []

        html = response.text
        soup = BeautifulSoup(html, "html.parser")
        full_links = []
        # Находим на странице все теги <a> с классом "but but-1" - в них присутствует ссылка на раздел
        links = soup.find_all("a", class_="but but-1")
        if links:
            for link in links:
                full_link = "https://stout.ru" + str(link.get('href'))
                full_links.append(full_link)

        return full_links

    except requests.exceptions.RequestException as e:  # Ловим любые ошибки с запросом
        print(f"Ошибка запроса: {e}")
        return []  # Возвращаем пустой список, чтобы продолжить выполнение программы


def section_decomposition(url):
    total_products = 0  # Счётчик товаров на странице

    try:
        response = requests.get(url)
        if response.status_code != 200:  # Проверяем код ответа
            print(f"Ошибка при запросе {url}. Статус код: {response.status_code}")
            return []

        html_content = response.text
        # Создание объекта BeautifulSoup для разбора HTML-структуры
        soup = BeautifulSoup(html_content, 'html.parser')

        # Извлекаем заголовок страницы (например, название раздела)
        page_title = soup.find('h1').get_text(strip=True)
        table_data = []

        # Находим все элементы товара с классом 'product-item'
        sku_list = soup.find_all('article', class_='product-item')
        #print(f"Найдено товаров: {len(sku_list)}")  # Проверяем количество найденных товаров

        for sku in sku_list:
            # Ищем элемент с артикулом
            sku_article_tag = sku.find("span", class_="product-item-sku a_pt_2")  # Учитываем только товары с классом 'a_pt_2'

            # Продолжаем только если нашли артикул с нужным классом
            if sku_article_tag:
                # Извлекаем артикул товара
                sku_article = sku_article_tag.get_text(strip=True)
                # Извлекаем название товара и ссылку на его страницу
                sku_name_tag = sku.find("a", class_="product-item-title")
                sku_name = sku_name_tag.get_text(strip=True)  # Название
                sku_link = "https://stout.ru" + str(sku_name_tag['href'])  # Ссылка на страницу товара

                # Извлекаем цену товара
                sku_price_tag = sku.find("strong", class_="block-title product-item-price")
                if sku_price_tag:
                    # Извлекаем только цифры из цены, игнорируя символ рубля
                    sku_price = ''.join([char for char in sku_price_tag.get_text(strip=True) if char.isdigit()])
                else:
                    sku_price = "Цена не указана"  # В случае отсутствия цены

                # Формируем результат и добавляем в таблицу
                sku_data = {
                    "Раздел": page_title,
                    "Артикул": sku_article,
                    "Название": sku_name,
                    "Цена": sku_price,
                    "Ссылка": sku_link,
                    "Дата": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                }
                table_data.append(sku_data)
                total_products += 1  # Счётчик товаров на странице

        # Проверяем результат
        if table_data:
            print_to_excel(table_data)
            return total_products  # Возвращаем количество найденных товаров
            #pprint.pprint(table_data)
        else:
            print("Товары не найдены или данные отсутствуют.")

    except requests.exceptions.RequestException as e:  # Ловим любые ошибки с запросом
        print(f"Ошибка запроса: {e}")
        return total_products  # Возвращаем количество найденных товаров


def get_total_pages(url):
    try:
        response = requests.get(url)
        if response.status_code != 200:  # Проверяем код ответа
            print(f"Ошибка при запросе {url}. Статус код: {response.status_code}")
            return 1  # Если не можем определить количество страниц, возвращаем хотя бы одну

        html_content = response.text
        soup = BeautifulSoup(html_content, 'html.parser')

        # Находим блок пагинации
        pagination = soup.find('div', class_='reviews-pagination')  # Блок пагинации

        if not pagination:
            print("Блок пагинации не найден. Возвращаем 1 страницу.")
            return 1  # Если пагинации нет, возвращаем 1 страницу

        # Находим все ссылки на страницы в блоке пагинации
        page_links = pagination.find_all('a', href=True)

        # Извлекаем номер последней страницы из href ссылки (например, "?PAGEN_1=3")
        total_pages = 1
        for link in page_links:
            match = re.search(r'PAGEN_\d+=(\d+)', link['href'])  # Ищем номер страницы в параметре PAGEN_
            if match:
                page_num = int(match.group(1))
                total_pages = max(total_pages, page_num)  # Находим максимальный номер страницы
        print(f"В текущем разделе {total_pages} страниц")
        return total_pages

    except requests.exceptions.RequestException as e:  # Ловим любые ошибки с запросом
        print(f"Ошибка запроса: {e}")
        return 1  # Возвращаем 1, если не можем получить количество страниц


def process_all_pages(base_url):
    total_products = 0  # Глобальный счётчик товаров
    total_pages = get_total_pages(base_url)  # Определяем количество страниц

    for page in range(1, total_pages + 1):
        # Формируем URL для каждой страницы, добавляя параметр пагинации
        page_url = f"{base_url}?PAGEN_1={page}"
        print(f"Обрабатываем страницу: {page_url}")

        # Вызываем функцию для обработки товаров на каждой странице
        products_on_page = section_decomposition(page_url)
        total_products += products_on_page  # Суммируем количество товаров со всех страниц

    print(f"Всего найдено товаров в разделе: {total_products}")
    return total_products


url = "https://www.stout.ru/catalog/"
# Начало отсчёта времени
start_time = time.time()
# Счётчик для подсчёта количества найденных товаров
total_products_found = 0

# Основной цикл парсинга
for x in get_urls_lvl_0(url):  # парсим корневой уровень каталога
    print(f"Прогон цикла ссылок 0 уровня: {x}")
    total_products_found +=process_all_pages(x)




# Конец отсчёта времени
end_time = time.time()

# Подсчитываем потраченное время
elapsed_time = end_time - start_time
elapsed_time_minutes = elapsed_time / 60  # переводим в минуты

# Вывод результатов
print("=========================================")
print(f"Программа завершена.")
print(f"Найдено товаров: {total_products_found}")
print(f"Время выполнения программы: {elapsed_time:.2f} секунд ({elapsed_time_minutes:.2f} минут)")
print("=========================================")