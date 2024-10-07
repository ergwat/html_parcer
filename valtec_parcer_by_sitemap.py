from bs4 import BeautifulSoup
import requests
from datetime import datetime
import pprint
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import os
import urllib.request
from urllib.request import urlopen
from bs4 import BeautifulSoup
import re

def table_decomposition(url):

    html_content = requests.get(url).text

    # Создание объекта BeautifulSoup
    soup = BeautifulSoup(html_content, 'html.parser')

    # Извлекаем заголовок страницы (тег <h1>)
    page_title = soup.find('h1').get_text(strip=True)

    # Находим заголовки таблицы
    headers = [th.get_text(strip=True) for th in soup.find_all('tr')[0].find_all('th')]

    # Инициализируем список для хранения записей
    table_data = []

    # Проходим по строкам таблицы (начиная со 2-й строки данных)
    for row in soup.find_all('tr')[2:]:
        # Извлекаем все ячейки в строке
        cells = row.find_all('td')
        # Проверяем, что это не строка с дополнительной информацией (например, строка с "*") и что количество ячеек совпадает с количеством заголовков
        if len(cells) == len(headers):
            # Создаем запись (словарь) для каждой строки таблицы
            row_data = {'Название': page_title}  # Добавляем заголовок страницы
            for i, cell in enumerate(cells):
                # Если ячейка содержит цену, убираем <em> и прочие теги
                if 'Цена' in headers[i]:
                    price_text = cell.get_text(strip=True).replace('p', '').strip()
                    row_data[headers[i]] = price_text
                else:
                    # Добавляем текст в соответствующий заголовок
                    row_data[headers[i]] = cell.get_text(strip=True)
            # Добавляем запись в список
            table_data.append(row_data)

    # Вывод результата
    #pprint.pprint(table_data)
    print_to_excel(table_data)

def print_to_excel(table_data):

    # Преобразование строковых значений цен в числа с плавающей запятой и округление до 2 знаков
    for row in table_data:
        # Преобразуем значение цены в float и округляем до 2 знаков
        row['Цена за единицу'] = round(float(row['Цена за единицу']), 2)

    # Получаем текущую дату и время в формате 'YYYY-MM-DD HH:MM:SS'
    current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Добавляем текущую дату и время в каждый ряд данных
    for row in table_data:
        row['Дата'] = current_datetime

    # Создание имени файла с текущей датой
    filename = f"competitors_parsing_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

    # Проверяем, существует ли файл
    if os.path.exists(filename):
        # Если файл существует, открываем его для добавления данных
        wb = load_workbook(filename)
        ws = wb.active
    else:
        # Если файл не существует, создаём новый
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"valtec_ru_{datetime.now().strftime('%Y-%m-%d')}"
        # Запись заголовков столбцов (это ключи первого словаря)
        headers = list(table_data[0].keys())  # теперь включая "Дата"
        ws.append(headers)

    # Определяем, с какой строки начинать добавление данных
    start_row = ws.max_row + 1

    # Запись данных
    for row in table_data:
        ws.append(list(row.values()))  # добавляем значения каждой строки

    # Сохранение файла
    wb.save(filename)

    print(f"Данные успешно сохранены в файл {filename}.")


def get_urls_lvl_2(url):
    response = requests.get(url)
    html = response.text
    soup = BeautifulSoup(html, "html.parser")
    #links = soup.find_all("a", class_="catalog2 subgroup")
    full_links = []
    # Находим div с классом "catalog2 subgroup"
    catalog_div = soup.find("div", class_="catalog2 subgroup")

    # Извлекаем все ссылки (теги <a>) внутри этого блока
    links = catalog_div.find_all("a")

    # Собираем полные ссылки
    for link in links:
        full_link = "https://valtec.ru" + str(link.get('href'))
        full_links.append(full_link)
        print(full_link)
    return full_links

''' ============== пытаемся выгрузить sitemap.xml ============='''
def find_all_product_links(url):
    """Функция для поиска всех ссылок на продукты."""
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    soup = BeautifulSoup(urllib.request.urlopen(req), 'html.parser')
    all_product_links = [link.get('href') for link in soup.find_all('loc') if
                         re.search('catalog/', link.text) ] #and link.text.endswith('.html')]
    print(all_product_links)
    return all_product_links


'''
url0 = "https://valtec.ru/catalog/"
url1 = "https://valtec.ru/catalog/filtry/"
url2 = "https://valtec.ru/catalog/filtry/filtry_kosye/"
url3 = "https://valtec.ru/catalog/filtry/filtry_kosye/filtruyushij_lement_dlya_kosyh_filtrov.html"

for i in get_urls_lvl_2(url2):
    print(i)
    table_decomposition(i)
'''

sitemap_url="https://valtec.ru/sitemap.xml"
#get_sitemap(sitemap_url)
for i in find_all_product_links(sitemap_url):
    table_decomposition(i)